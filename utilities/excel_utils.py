import openpyxl

def get_data_from_excel(file_path, sheet_name="LoginData"):
    workbook = openpyxl.load_workbook(file_path)
    
    # ఒకవేళ ఇచ్చిన సీట్ నేమ్ దొరకకపోతే, ఎక్సెల్‌లో ఉన్న ఫస్ట్ షీట్‌ని ఆటోమేటిక్‌గా సెలెక్ట్ చేసుకుంటుంది
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
        
    data = []
    
    # 2nd Row నుండి హెడర్స్ కాకుండా మిగతా డేటాని రీడ్ చేస్తుంది
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # ఖాళీ రోస్ (empty rows) ఉంటే వదిలేస్తుంది
            data.append(list(row))
            
    return data